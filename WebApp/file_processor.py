# import os
# import pandas as pd
# import zipfile
# from openpyxl import load_workbook
# from datetime import datetime

# currency_cols = ['Split Amount', 'Override', 'Override By Carrier', 'Split Amount By Carrier']

# def process_file_with_columns(input_path: str, output_path: str, selected_columns: list):
#     ext = os.path.splitext(input_path)[1].lower()
#     if ext == '.csv':
#         df = pd.read_csv(input_path)
#     elif ext in ('.xls', '.xlsx'):
#         try:
#             df = pd.read_excel(input_path, engine='openpyxl')
#         except (zipfile.BadZipFile, ValueError):
#             df = pd.read_csv(input_path)
#     else:
#         raise ValueError(f"Unsupported file type: {ext}")

#     # Parse 'Statement Date'
#     if 'Statement Date' in df.columns:
#         df['Statement Date'] = pd.to_datetime(df['Statement Date'], errors='coerce').dt.normalize()
#         df.sort_values(by='Statement Date', inplace=True)

#     # Clean currency columns
#     for col in currency_cols:
#         if col in df.columns:
#             df[col] = df[col].astype(str).str.replace(r'[\\$,]', '', regex=True).pipe(pd.to_numeric, errors='coerce')

#     # Convert percent
#     if 'Override Percent' in df.columns:
#         df['Override Percent'] = df['Override Percent'].astype(str).str.rstrip('%').str.strip().pipe(pd.to_numeric, errors='coerce').div(100)

#     # Exclude renewals
#     if 'Product Option' in df.columns:
#         df = df[~df['Product Option'].astype(str).str.contains('renewal', case=False, na=False)]

#     # Filter only selected columns that exist
#     available = [col for col in selected_columns if col in df.columns]
#     filtered_df = df[available]
#     filtered_df.to_excel(output_path, index=False, engine='openpyxl')

#     wb = load_workbook(output_path)
#     ws = wb.active
#     ws.title = 'Filtered Data'
#     col_idx = {name: idx + 1 for idx, name in enumerate(filtered_df.columns)}

#     for row in range(2, ws.max_row + 1):
#         for cur in currency_cols:
#             if cur in col_idx:
#                 ws.cell(row, col_idx[cur]).number_format = '$#,##0.00'
#         if 'Override Percent' in col_idx:
#             ws.cell(row, col_idx['Override Percent']).number_format = '0.00%'
#         if 'Statement Date' in col_idx:
#             ws.cell(row, col_idx['Statement Date']).number_format = 'MM/DD/YYYY'

#     wb.save(output_path)

#     # Summary only if date exists
#     if 'Statement Date' in df.columns:
#         df['Year'] = df['Statement Date'].dt.year
#         most_recent_year = df['Year'].max()
#         df = df[df['Year'] >= most_recent_year - 4]

#     summary = (
#         df.groupby(['Carrier', 'Year'], dropna=False)['Split Amount']
#         .sum()
#         .reset_index()
#         .sort_values(['Carrier', 'Year'])
#     )

#     running_balance = 0
#     ws_summary = wb.create_sheet(title='Yearly Summary')
#     ws_summary.append(['Carrier', 'Year', 'Total Split Amount', 'Reimbursement Account', 'Reimbursed Amount', 'Balance'])
#     for carrier, year, total in summary.itertuples(index=False, name=None):
#         reimb = total * 0.001
#         running_balance += reimb
#         ws_summary.append([carrier, year, total, reimb, "N/A", running_balance])

#     for r in range(2, ws_summary.max_row + 1):
#         ws_summary.cell(row=r, column=3).number_format = '$#,##0.00'
#         ws_summary.cell(row=r, column=4).number_format = '$#0.00'
#         ws_summary.cell(row=r, column=6).number_format = '$#0.00'

#     wb.save(output_path)



import os
import pandas as pd
import zipfile
from openpyxl import load_workbook
from datetime import datetime

currency_cols = ['Split Amount', 'Override', 'Override By Carrier', 'Split Amount By Carrier']

def process_file_with_columns(input_path: str, output_path: str, selected_columns: list):
    ext = os.path.splitext(input_path)[1].lower()
    if ext == '.csv':
        df = pd.read_csv(input_path)
    elif ext in ('.xls', '.xlsx'):
        try:
            df = pd.read_excel(input_path, engine='openpyxl')
        except (zipfile.BadZipFile, ValueError):
            df = pd.read_csv(input_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    if 'Statement Date' in df.columns:
        df['Statement Date'] = pd.to_datetime(df['Statement Date'], errors='coerce').dt.normalize()
        df.sort_values(by='Statement Date', inplace=True)

    for col in currency_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(r'[\\$,]', '', regex=True).pipe(pd.to_numeric, errors='coerce')

    if 'Override Percent' in df.columns:
        df['Override Percent'] = df['Override Percent'].astype(str).str.rstrip('%').str.strip().pipe(pd.to_numeric, errors='coerce').div(100)

    if 'Product Option' in df.columns:
        df = df[~df['Product Option'].astype(str).str.contains('renewal', case=False, na=False)]

    available = [col for col in selected_columns if col in df.columns]
    filtered_df = df[available]
    filtered_df.to_excel(output_path, index=False, engine='openpyxl')

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = 'Filtered Data'
    col_idx = {name: idx + 1 for idx, name in enumerate(filtered_df.columns)}

    for row in range(2, ws.max_row + 1):
        for cur in currency_cols:
            if cur in col_idx:
                ws.cell(row, col_idx[cur]).number_format = '$#,##0.00'
        if 'Override Percent' in col_idx:
            ws.cell(row, col_idx['Override Percent']).number_format = '0.00%'
        if 'Statement Date' in col_idx:
            ws.cell(row, col_idx['Statement Date']).number_format = 'MM/DD/YYYY'

    wb.save(output_path)

    if 'Statement Date' in df.columns:
        df['Year'] = df['Statement Date'].dt.year
        most_recent_year = df['Year'].max()
        df = df[df['Year'] >= most_recent_year - 4]

    summary = (
        df.groupby(['Carrier', 'Year'], dropna=False)['Split Amount']
        .sum()
        .reset_index()
        .sort_values(['Carrier', 'Year'])
    )

    first = True
    running_balance = 0
    ws_summary = wb.create_sheet(title='Yearly Summary')
    ws_summary.append(['Carrier', 'Year', 'Total Split Amount', 'Reimbursement Account', 'Reimbursed Amount', 'Balance'])

    for carrier, year, total in summary.itertuples(index=False, name=None):
        reimb = total * 0.001
        row_num = ws_summary.max_row + 1
        running_balance += reimb
        if first:
            ws_summary.append([carrier, year, total, reimb, "N/A", reimb])
            first = False
        else:
            ws_summary.append([carrier, year, total, reimb, "N/A", None])
            formula = f"=F{row_num - 1} + D{row_num} - IF(ISNUMBER(E{row_num}), E{row_num}, 0)"
            ws_summary.cell(row=row_num, column=6).value = formula



    for r in range(2, ws_summary.max_row + 1):
        ws_summary.cell(row=r, column=3).number_format = '$#,##0.00'
        ws_summary.cell(row=r, column=4).number_format = '$#0.00'
        ws_summary.cell(row=r, column=6).number_format = '$#0.00'

    wb.save(output_path)